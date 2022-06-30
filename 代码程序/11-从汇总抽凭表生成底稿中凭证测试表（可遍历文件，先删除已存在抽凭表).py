import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import glob
title_font = Font(name='黑体', size=14, bold=True)
content_font = Font(name='宋体', size=10, bold=False)
thin = Side(border_style="thin")
border = Border(top=thin, left=thin, right=thin, bottom=thin)
center_alignment=Alignment(horizontal='center',vertical='center')
filearray=[]
filelocation=glob.glob(r"C:\Users\DELL\Desktop\210731底稿-非最新\*.xlsx")  #原始底稿存放位置，建议操作之前做好备份
wb2=load_workbook(r"C:\Users\DELL\Desktop\项目文件d\凭证更新\11-抽凭表模板.xlsx")#抽凭表位置，需按照模板填写
ws2=wb2["Sheet1"]
Maxrow=ws2.max_row
namelist=[]

for filename in filelocation:
    filearray.append(filename)

    list1=filename.split("\\")
    print(list1)
    name1=list1[5] #取Excel名称所在的位置，如C:\Users\DELL\Desktop\210731底稿-非最新\。xlsx，中，C为0，User为1，底稿名称位置是最后.xlsxd的位置
    name2=name1[:-10] #从Excel名称中保留科目名称，这个数字是截取掉的字符的长度，比如_邦克实业.xlsx,共十个字符
    print(name2)
    n=1
    ifcontains=False
    while n <Maxrow+1:
        if ws2.cell(n,1).value==name2:
            ifcontains=True
        n=n+1
    print(ifcontains)
    print(name2)
    if name2.find("~$") < 0:

        wb1=load_workbook(filename)
        #删除已存在的抽凭表
        sheets=wb1.sheetnames
        for N in range(len(sheets)):
            if sheets[N].find("细节测试") >= 0 or sheets[N].find("凭证测试") >= 0:
                ws=wb1[sheets[N]]
                wb1.remove(ws)

        if ifcontains:
            #wb1=load_workbook(filename)
            ws = wb1.create_sheet("记账凭证测试表")
            ws.cell(4,1).value="                                          记账凭证测试表(通用式)"
            ws.cell(4,1).font=title_font
            ws.column_dimensions['G'].width =18
            ws.column_dimensions['H'].width =18
            ws.column_dimensions['B'].width =12
            m=3
            while m<16:
                ws.cell(5,m-2).value=ws2.cell(1,m).value
                ws.cell(5, m-2).font=content_font
                ws.cell(5, m-2).border=border
                ws.cell(5, m-2).alignment=center_alignment
                m=m+1
            i=1
            j=6
            xh=1
            while i <Maxrow+1:

                if ws2.cell(i,1).value==name2:

                    ws.cell(j, 1).value=xh
                    ws.cell(j, 1).font=Font(name='arial', size=10, bold=False)
                    ws.cell(j, 1).border=border
                    ws.cell(j, 1).alignment=center_alignment
                    # 日期一列格式
                    ws.cell(j, 2).value=ws2.cell(i, 4).value
                    ws.cell(j, 2).font=content_font
                    ws.cell(j, 2).border=border
                    ws.cell(j, 2).alignment=center_alignment
                    ws.cell(j, 2).number_format="mm-dd-yy"
                    #摘要，明细科目及对方科目格式
                    m=5
                    while m<9:
                        ws.cell(j,m-2).value=ws2.cell(i,m).value
                        ws.cell(j, m-2).font=content_font
                        ws.cell(j, m-2).border=border
                        m=m+1
                    #设置借方金额格式
                    ws.cell(j, 7).value=ws2.cell(i, 9).value
                    ws.cell(j, 7).font=Font(name='arial', size=10, bold=False)
                    ws.cell(j, 7).border=border
                    ws.cell(j, 7).number_format="_ * #,##0.00_ "
                    # 设置贷方金额格式
                    ws.cell(j, 8).value=ws2.cell(i, 10).value
                    ws.cell(j, 8).font=Font(name='arial', size=10, bold=False)
                    ws.cell(j, 8).border=border
                    ws.cell(j, 8).number_format="_ * #,##0.00_ "
                    #设置√号四列格式
                    m=11
                    while m < 14:
                        ws.cell(j, m-2).value=ws2.cell(i, m).value
                        ws.cell(j, m-2).font=content_font
                        ws.cell(j, m-2).border=border
                        ws.cell(j, m-2).alignment=center_alignment
                        m=m + 1
                    #设置附件一列格式
                    ws.cell(j, 12).value=ws2.cell(i, 14).value
                    ws.cell(j, 12).font=content_font
                    ws.cell(j, 12).border=border
                    #设置索引号列格式
                    ws.cell(j, 13).value=ws2.cell(i, 15).value
                    ws.cell(j, 13).font=Font(name='arial', size=10, bold=False)
                    ws.cell(j, 13).border=border

                    j=j+1
                    xh=xh+1

                    print(i)
                i=i+1
            ws.cell(j+3, 1).value="审计结论："
            ws.cell(j+3, 1).font=Font(name='宋体', size=10, bold=True)
            ws.cell(j+4, 1).value="    经审计，未见异常"
            ws.cell(j+4, 1).font=Font(name='宋体', size=10, bold=True)
        wb1.save(r"C:\Users\DELL\Desktop\修改后底稿\{}".format(name1))




