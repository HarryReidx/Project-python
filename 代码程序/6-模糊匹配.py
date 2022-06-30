from fuzzywuzzy import fuzz
from fuzzywuzzy import process
from openpyxl import load_workbook
wb=load_workbook(r'C:\Users\os\Desktop\模糊匹配.xlsx')
#获取workbook中所有的表格
sheets = wb.sheetnames
sheet=wb[sheets[0]]
m=1
while m<250:
    j=1
    sheet.cell(m, 2).value=sheet.cell(1, 4).value
    b=0
    while j<1000:
        str1=str(sheet.cell(m, 1).value)
        str2=str(sheet.cell(j, 4).value)
        a=fuzz.ratio(str1.strip("有限责任公司").strip("有限公司").strip("公司"),str2.strip("有限责任公司").strip("有限公司").strip("公司"))
        if a>b:
            b=a
            sheet.cell(m, 2).value=sheet.cell(j, 4).value
            sheet.cell(m, 3).value=b
        j=j+1
    print(m)
    m=m+1

wb.save(r'C:\Users\os\Desktop\模糊匹配后文件.xlsx')




# print(fuzz.ratio("北钞","北京印钞有限公司")) #简单匹配
# print(fuzz.partial_ratio("北钞","北京印钞有限公司"))  #非完全匹配
# print(fuzz.token_sort_ratio("北钞","北京印钞有限公司")) #忽略顺序匹配
# print(fuzz.token_set_ratio("北钞","北京印钞有限公司")) #去除重集匹配


